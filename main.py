import openpyxl
import json
from datetime import datetime

def convert_excel_to_json(template_file_path, data_file_path, output_file_path):
    # テンプレートファイルを読み込み
    template_workbook = openpyxl.load_workbook(template_file_path)
    template_sheet = template_workbook.active

    # 取り込み対象ファイルを読み込み
    data_workbook = openpyxl.load_workbook(data_file_path, data_only=True)  # data_only=Trueでセルの数式を評価せずに値を読み込む
    data_sheet = data_workbook.active

    def process_key(json_data, key, data_cell_value):
        keys = key.split(".")
        current_data = json_data

        for k in keys[:-1]:
            if k not in current_data:
                current_data[k] = {}
            current_data = current_data[k]

        current_data[keys[-1]] = data_cell_value

    # JSONデータを格納するための辞書を作成
    json_data = {}

    # テンプレートファイルのセルと取り込み対象ファイルのセルを対応づけてJSONデータを作成
    for template_row, data_row in zip(template_sheet.iter_rows(), data_sheet.iter_rows()):
        for template_cell, data_cell in zip(template_row, data_row):
            # テンプレートファイルのセルのフォント情報を取得
            font = template_cell.font
            if font.color.rgb == "FFFF0000":  # 赤色の場合
                key = template_cell.value
                
                # if isinstance(data_cell.value, datetime):
                #     # 日付データをISO 8601フォーマットの文字列に変換
                #     data_cell.value = data_cell.value.isoformat()

                if "." in key:
                    process_key(json_data, key, str(data_cell.value))  # セルの値を文字列に変換
                else:
                    json_data[key] = str(data_cell.value)  # セルの値を文字列に変換

    # JSONデータをUTF-8エンコードでファイルに書き出す
    with open(output_file_path, "w", encoding="utf-8") as json_file:
        json.dump(json_data, json_file, ensure_ascii=False, indent=4)

    print(f"JSONファイルがUTF-8エンコードで {output_file_path} に作成されました.")

# 関数をテスト
template_file = "template.xlsx"
data_file = "data.xlsx"
output_file = "output.json"

convert_excel_to_json(template_file, data_file, output_file)
